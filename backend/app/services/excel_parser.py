"""
Excel parser for bulk supplier shipment upload.
Returns parsed rows + per-row validation errors before import.
"""
import io
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


EXPECTED_COLUMNS = [
    "supplier_name", "supplier_country", "factory_name", "contact_email",
    "cn_code", "production_quantity", "unit", "electricity_consumption_mwh",
    "fuel_type", "fuel_consumption_gj", "calculation_method",
    "shipment_date", "invoice_number", "reporting_period",
]

VALID_UNITS = ["tonne", "MWh", "m3"]
VALID_FUEL_TYPES = ["coal", "natural_gas", "diesel", "heavy_fuel_oil", "lpg", "none"]
VALID_METHODS = ["actual", "default"]


def generate_template() -> bytes:
    """Generate downloadable Excel template with validation dropdowns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = EXPECTED_COLUMNS
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Data validation dropdowns
    unit_dv = DataValidation(type="list", formula1='"tonne,MWh,m3"', allow_blank=True)
    fuel_dv = DataValidation(type="list", formula1='"coal,natural_gas,diesel,heavy_fuel_oil,lpg,none"', allow_blank=True)
    method_dv = DataValidation(type="list", formula1='"actual,default"', allow_blank=True)

    ws.add_data_validation(unit_dv)
    ws.add_data_validation(fuel_dv)
    ws.add_data_validation(method_dv)

    unit_col = headers.index("unit") + 1
    fuel_col = headers.index("fuel_type") + 1
    method_col = headers.index("calculation_method") + 1

    for row in range(2, 1002):
        unit_dv.add(ws.cell(row=row, column=unit_col))
        fuel_dv.add(ws.cell(row=row, column=fuel_col))
        method_dv.add(ws.cell(row=row, column=method_col))

    # Sample row
    sample = [
        "Steel Co LLC", "ARE", "Dubai Steel Factory", "contact@steelco.ae",
        "7208", "1000", "tonne", "250", "coal", "500", "default",
        "2026-01-15", "INV-001", "2026-Q1",
    ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    Parse uploaded Excel file.
    Returns: { "rows": [...], "errors": { row_num: [error_msg] }, "valid_count": int }
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        return {"rows": [], "errors": {"header": [f"Missing columns: {missing}"]}, "valid_count": 0}

    col_idx = {col: headers.index(col) for col in EXPECTED_COLUMNS}

    rows = []
    errors: dict[int, list[str]] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(col: str) -> Any:
            idx = col_idx[col]
            return row[idx] if idx < len(row) else None

        row_errors = []
        row_data: dict[str, Any] = {}

        for col in EXPECTED_COLUMNS:
            row_data[col] = get(col)

        if not row_data["supplier_name"]:
            row_errors.append("supplier_name is required")
        if not row_data["supplier_country"] or len(str(row_data["supplier_country"])) != 3:
            row_errors.append("supplier_country must be ISO 3-letter code")
        if not row_data["cn_code"]:
            row_errors.append("cn_code is required")

        try:
            qty = float(row_data["production_quantity"] or 0)
            if qty <= 0:
                row_errors.append("production_quantity must be > 0")
            row_data["production_quantity"] = qty
        except (TypeError, ValueError):
            row_errors.append("production_quantity must be a number")

        if str(row_data.get("unit", "")).lower() not in VALID_UNITS:
            row_errors.append(f"unit must be one of: {VALID_UNITS}")
        if str(row_data.get("fuel_type", "")).lower() not in VALID_FUEL_TYPES:
            row_errors.append(f"fuel_type must be one of: {VALID_FUEL_TYPES}")
        if str(row_data.get("calculation_method", "")).lower() not in VALID_METHODS:
            row_errors.append(f"calculation_method must be one of: {VALID_METHODS}")

        if row_errors:
            errors[row_num] = row_errors
        else:
            rows.append(row_data)

    return {
        "rows": rows,
        "errors": errors,
        "valid_count": len(rows),
        "error_count": len(errors),
        "total_rows": len(rows) + len(errors),
    }
